import openpyxl
import json
from pathlib import Path

excel_path = Path("app/zen_rules/Bank_Eligibility_Matrix_v1.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb["decision table"]

headers = [str(sheet.cell(1, c).value).strip() if sheet.cell(1, c).value is not None else f"Col{c}" for c in range(1, sheet.max_column + 1)]
print(f"Total Columns: {len(headers)}")

matrix = {}
for r in range(2, sheet.max_row + 1):
    bank_name = str(sheet.cell(r, 1).value).strip()
    row_data = {}
    for c in range(1, sheet.max_column + 1):
        h = headers[c - 1]
        val = sheet.cell(r, c).value
        row_data[h] = val
    matrix[bank_name] = row_data

with open("scratch/matrix_dump.json", "w", encoding="utf-8") as f:
    json.dump({"headers": headers, "matrix": matrix}, f, indent=2, default=str)

print("Matrix dump saved to scratch/matrix_dump.json")
for bank, data in matrix.items():
    print(f"\n--- {bank} ---")
    for k, v in list(data.items())[:15]:
        print(f"  {k}: {v}")
