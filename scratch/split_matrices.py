"""Split Bank_Eligibility_Matrix_v1.xlsx into two entity-scoped views.

Column ownership is derived from what app/services/bre_engine.py actually
evaluates per entity type, not from column names. Notably, a Company applicant
is scored on the SELF-EMPLOYED income rules (the engine maps Company ->
occupation "Self-Employed"), and never reaches the address, NRI, salaried or
co-applicant columns.

The 24 universal + shared columns appear in BOTH files by necessity — a CIBIL
floor applies to a company as much as to an individual. That duplication is the
split's main hazard, so verify_overlap() re-reads both outputs and asserts the
shared cells are byte-identical.
"""

import shutil
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "app" / "zen_rules" / "Bank_Eligibility_Matrix_v1.xlsx"
SHEET = "decision table"
OUT_INDIVIDUAL = ROOT / "app" / "zen_rules" / "bank_Individual_Eligibility_Matrix.xlsx"
OUT_COMPANY = ROOT / "app" / "zen_rules" / "bank_Company_Organization_Eligibility_Matrix.xlsx"

# Identity + policy that binds every applicant regardless of entity type.
UNIVERSAL = [
    0, 1,                          # Bank Name, Description
    2,                             # CIBIL Score
    3, 4, 5, 6, 7, 8, 9, 10,       # Write-offs (7 product classes + CC cap)
    11, 12, 13,                    # DPD, Loan enquiry, Currently Outstanding
    17, 18,                        # Existing A/C Holder, Existing Car Loan
    48,                            # Business Proof
    53,                            # EMI / Income Ratio
]

# Income rules shared by Individual-SE, HUF and Company.
SELF_EMPLOYED_SHARED = [42, 43, 44, 45, 46, 47]

# Natural-person rules: demographics, residence/guarantor, NRI, employment,
# salary, rental income, sole-trader entity types, Form 16, co-applicant.
INDIVIDUAL_ONLY = [
    14, 15, 16,
    19, 20, 21, 22, 23, 24,
    25, 26, 27,
    28, 29, 30, 31, 32, 33, 34, 35, 36, 37,
    38, 39, 40, 41,
    49, 50,
    55,
    56, 57, 58, 59, 60, 61,
]

HUF_ONLY = [54]
CORPORATE_ENTITY_TYPES = [51, 52]  # Private Limited, Public Limited

INDIVIDUAL_COLUMNS = sorted(UNIVERSAL + SELF_EMPLOYED_SHARED + INDIVIDUAL_ONLY + HUF_ONLY)
COMPANY_COLUMNS = sorted(UNIVERSAL + SELF_EMPLOYED_SHARED + CORPORATE_ENTITY_TYPES)


def write_subset(rows: list[tuple], columns: list[int], destination: Path, title: str) -> None:
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = SHEET
    for row in rows:
        sheet.append([row[i] for i in columns])
    sheet.freeze_panes = "C2"
    book.save(destination)
    print(f"  {destination.name:<52} {len(columns):>2} cols x {len(rows) - 1} banks  [{title}]")


def verify_overlap() -> int:
    """The shared columns exist twice now; prove the copies agree."""
    ind = list(openpyxl.load_workbook(OUT_INDIVIDUAL, data_only=True)[SHEET].iter_rows(values_only=True))
    com = list(openpyxl.load_workbook(OUT_COMPANY, data_only=True)[SHEET].iter_rows(values_only=True))
    ind_index = {h: i for i, h in enumerate(ind[0])}
    com_index = {h: i for i, h in enumerate(com[0])}
    shared = set(ind_index) & set(com_index)

    mismatches = 0
    for ind_row, com_row in zip(ind[1:], com[1:]):
        for header in shared:
            a, b = ind_row[ind_index[header]], com_row[com_index[header]]
            if a != b:
                print(f"  DRIFT {ind_row[0]}.{header}: individual={a!r} company={b!r}")
                mismatches += 1
    print(f"  overlap: {len(shared)} shared columns x {len(ind) - 1} banks, {mismatches} mismatches")
    return mismatches


def main() -> int:
    rows = list(openpyxl.load_workbook(SOURCE, data_only=True)[SHEET].iter_rows(values_only=True))
    header = rows[0]

    print(f"source: {SOURCE.name} ({len(header)} cols x {len(rows) - 1} banks)\n")
    write_subset(rows, INDIVIDUAL_COLUMNS, OUT_INDIVIDUAL, "Individual + HUF")
    write_subset(rows, COMPANY_COLUMNS, OUT_COMPANY, "Company / Organization")

    dropped = sorted(set(range(len(header))) - set(INDIVIDUAL_COLUMNS) - set(COMPANY_COLUMNS))
    print(f"\n  columns in neither output: {[header[i] for i in dropped] or 'none'}")
    print()
    return verify_overlap()


if __name__ == "__main__":
    sys.exit(1 if main() else 0)
