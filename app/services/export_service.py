"""Binary report generation for onboarding evaluations.

Both builders assemble entirely in a `BytesIO` — no temp files, no disk writes
on the request path. Applicant identity is rendered from the already-masked
values persisted on the application row (PAN reaches the database only as
`AB******4F`), so an exported document cannot leak what the audit trail does
not hold.
"""

from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

# Verdict palette, shared by both formats so the documents read as one system.
PASS_FILL = "D9F2E3"
FAIL_FILL = "FBD9D9"
HEADER_FILL = "1E293B"
PASS_RGB = colors.HexColor("#15803D")
FAIL_RGB = colors.HexColor("#B91C1C")
INK_RGB = colors.HexColor("#0F172A")

_THIN = Side(style="thin", color="D0D5DD")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _rows_for_bank(report: Dict[str, Any]) -> List[Tuple[str, str, str, str, str, str]]:
    """Flatten one bank's report into (status, rule_id, name, value, limit, note)."""
    rows = [
        (r["status"], r["rule_id"], r["parameter_name"], r["user_value"],
         r["limit_value"], r["description"])
        for r in report.get("passed_rules", [])
    ]
    rows += [
        (r["status"], r["rule_id"], r["parameter_name"], r["user_value"],
         r["limit_value"], r["description"])
        for r in report.get("failed_rules", [])
    ]
    return rows


def build_excel(
    application: Dict[str, Any], evaluation_report: Dict[str, Any]
) -> BytesIO:
    """One summary sheet plus a per-bank sheet, PASS rows green, FAIL rows red."""
    book = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)

    def write_header(sheet, labels: List[str]) -> None:
        sheet.append(labels)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER

    def autosize(sheet, widths: List[int]) -> None:
        for i, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = width

    # --- Summary ------------------------------------------------------------
    summary = book.active
    summary.title = "Summary"
    write_header(summary, ["Field", "Value"])
    for label, value in application.items():
        summary.append([label, "" if value is None else str(value)])
    autosize(summary, [34, 52])

    summary.append([])
    summary.append(["Bank", "Eligible", "Passed", "Failed"])
    for cell in summary[summary.max_row]:
        cell.font = Font(bold=True)
        cell.border = _BORDER
    for bank, report in evaluation_report.items():
        eligible = bool(report.get("is_eligible"))
        summary.append([
            bank, "YES" if eligible else "NO",
            len(report.get("passed_rules", [])), len(report.get("failed_rules", [])),
        ])
        fill = PatternFill("solid", fgColor=PASS_FILL if eligible else FAIL_FILL)
        for cell in summary[summary.max_row]:
            cell.fill = fill
            cell.border = _BORDER

    # --- One sheet per bank -------------------------------------------------
    for bank, report in evaluation_report.items():
        sheet = book.create_sheet(title=bank[:31])
        write_header(sheet, ["Status", "Rule ID", "Rule", "Applicant Value", "Bank Limit", "Reason"])
        for status, rule_id, name, value, limit, note in _rows_for_bank(report):
            sheet.append([status, rule_id, name, value, limit, note])
            fill = PatternFill("solid", fgColor=PASS_FILL if status == "PASS" else FAIL_FILL)
            for cell in sheet[sheet.max_row]:
                cell.fill = fill
                cell.border = _BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.freeze_panes = "A2"
        autosize(sheet, [10, 12, 30, 22, 22, 60])

    stream = BytesIO()
    book.save(stream)
    stream.seek(0)
    return stream


def build_pdf(
    application: Dict[str, Any], evaluation_report: Dict[str, Any]
) -> BytesIO:
    """A4 report: applicant parameters, then a per-bank rule table."""
    stream = BytesIO()
    doc = SimpleDocTemplate(
        stream, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm, topMargin=16 * mm, bottomMargin=16 * mm,
        title="FlowBRE Eligibility Report",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=16, textColor=INK_RGB, spaceAfter=2)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=12, textColor=INK_RGB, spaceBefore=10)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=7.5, leading=9.5)

    story: List[Any] = [
        Paragraph("FlowBRE — Loan Eligibility Report", h1),
        Paragraph("Evaluated against the partner bank policy matrix.", styles["Normal"]),
        Spacer(1, 8),
        Paragraph("Applicant Parameters", h2),
    ]

    param_rows = [["Field", "Value"]] + [
        [k, Paragraph("" if v is None else str(v), small)] for k, v in application.items()
    ]
    params = Table(param_rows, colWidths=[55 * mm, 119 * mm], repeatRows=1)
    params.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), INK_RGB),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D5DD")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]))
    story.append(params)

    for bank, report in evaluation_report.items():
        eligible = bool(report.get("is_eligible"))
        verdict = "ELIGIBLE" if eligible else "NOT ELIGIBLE"
        heading = Paragraph(
            f'{bank} — <font color="{"#15803D" if eligible else "#B91C1C"}">{verdict}</font>', h2
        )

        data = [["", "Rule ID", "Rule", "Value", "Limit", "Reason"]]
        styling = [
            ("BACKGROUND", (0, 0), (-1, 0), INK_RGB),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D5DD")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]
        for index, (status, rule_id, name, value, limit, note) in enumerate(
            _rows_for_bank(report), start=1
        ):
            passed = status == "PASS"
            data.append([
                "PASS" if passed else "FAIL", rule_id,
                Paragraph(name, small), Paragraph(value, small),
                Paragraph(limit, small), Paragraph(note, small),
            ])
            styling.append((
                "BACKGROUND", (0, index), (-1, index),
                colors.HexColor("#EAF7EF" if passed else "#FDECEC"),
            ))
            styling.append(("TEXTCOLOR", (0, index), (0, index), PASS_RGB if passed else FAIL_RGB))
            styling.append(("FONTNAME", (0, index), (0, index), "Helvetica-Bold"))

        table = Table(data, colWidths=[12 * mm, 18 * mm, 34 * mm, 26 * mm, 26 * mm, 58 * mm], repeatRows=1)
        table.setStyle(TableStyle(styling))
        # Keep a bank heading with at least the start of its table.
        story.append(KeepTogether([heading, table]))

    doc.build(story)
    stream.seek(0)
    return stream
