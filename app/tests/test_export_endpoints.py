"""Coverage for the evaluation audit report and its PDF / Excel exports."""

import asyncio
import zipfile
from datetime import datetime, timezone
from io import BytesIO
from typing import Any
from unittest.mock import AsyncMock, MagicMock

import openpyxl
import pytest
from fastapi.testclient import TestClient

import app.core.redis as redis_module
from app.api.deps import get_db, get_redis
from app.db.models.application import ApplicationModel
from app.db.models.rule_execution import RuleExecutionModel
from app.main import app
from app.services.bre_engine import bre_engine_service
from app.services.export_service import build_excel, build_pdf

APPLICATION_ID = "app-export-1"

EVALUATION = asyncio.run(
    bre_engine_service.evaluate_application(
        {
            "occupation": "Salaried", "age": 34, "selected_bank": "BOI",
            "existing_account_bank": "BOI", "net_monthly_salary": 20000,
            "credit_bureau": {"cibil_score": 650, "dpd_history": [5], "write_off_amount": 0},
        }
    )
)

APP_ROW = ApplicationModel(
    id=APPLICATION_ID, tenant_id="tenant_beta", entity_type="Individual",
    applicant_name="Rohan Sharma", pan_masked="AB******4F", selected_bank="BOI",
    loan_type="Auto Loan", profile_type="Salaried", occupation="Salaried",
    property_status="OWNED", pincode="560001", cibil_score=650, max_dpd_days=5,
    loan_enquiry_count=0, currently_outstanding=0.0, write_off_amount=0.0,
    business_itr_years=None, status=EVALUATION["status"],
    created_at=datetime.now(timezone.utc),
)
EXEC_ROW = RuleExecutionModel(
    application_id=APPLICATION_ID, tenant_id="tenant_beta", bank_code="BOI",
    eligible=False, execution_time_ms=1.0,
    evaluation_report_json=EVALUATION["evaluation_report"],
)


def _result(value: Any) -> MagicMock:
    result = MagicMock()
    result.scalar_one_or_none.return_value = value
    return result


# Two SELECTs per export: the application row, then its rule execution.
rows_to_return: list[Any] = []


async def mock_get_db():
    session = MagicMock()

    async def execute(*_args, **_kwargs):
        return _result(rows_to_return.pop(0)) if rows_to_return else _result(None)

    session.execute = execute
    future: Any = asyncio.Future()
    future.set_result(MagicMock())
    session.flush.return_value = future
    session.commit.return_value = future
    session.rollback.return_value = future
    yield session


mock_redis = AsyncMock()
mock_redis.incr.return_value = 1
mock_redis.get.return_value = None
redis_module.redis_client = mock_redis


async def mock_get_redis():
    return mock_redis


client = TestClient(app)


@pytest.fixture(autouse=True)
def _override_dependencies():
    """Install this module's DB mock per-test.

    `app.dependency_overrides` is global and every test module assigns it at
    import time, so the last import silently wins for the whole session. Scoping
    the override to each test makes these cases independent of import order.
    """
    previous = dict(app.dependency_overrides)
    app.dependency_overrides[get_db] = mock_get_db
    app.dependency_overrides[get_redis] = mock_get_redis
    yield
    app.dependency_overrides.clear()
    app.dependency_overrides.update(previous)


def _stage(app_row: Any = APP_ROW, exec_row: Any = EXEC_ROW) -> None:
    rows_to_return.clear()
    # set_tenant_rls_context issues its own execute first.
    rows_to_return.extend([None, app_row, exec_row])


# --- Report payload -------------------------------------------------------- #


def test_evaluation_report_covers_every_bank() -> None:
    report = EVALUATION["evaluation_report"]
    assert set(report) == set(EVALUATION["bank_eligibility"])
    for bank, entry in report.items():
        assert entry["is_eligible"] == EVALUATION["bank_eligibility"][bank]
        assert entry["passed_rules"] or entry["failed_rules"]


def test_report_rows_carry_value_and_limit() -> None:
    failed = EVALUATION["evaluation_report"]["BOI"]["failed_rules"]
    cibil = next(r for r in failed if r["rule_id"] == "BUR-405")
    assert cibil["parameter_name"] == "CIBIL Score"
    assert cibil["status"] == "FAIL"
    # BOI publishes a personal-loan floor as well, so its limit names both and
    # the value stays the bare score while no PL score was supplied (add-on §1).
    assert cibil["user_value"] == "650"
    assert cibil["limit_value"] == ">= 701 or PL >= 701"
    assert cibil["description"]


def test_every_row_explains_itself() -> None:
    """The audit grid is only useful if a PASS states its outcome too."""
    for code, report in EVALUATION["evaluation_report"].items():
        for row in report["passed_rules"] + report["failed_rules"]:
            assert row["description"], f"{code}/{row['rule_id']} has no description"
            assert row["status"] in ("PASS", "FAIL")
        assert all(r["status"] == "PASS" for r in report["passed_rules"])
        assert all(r["status"] == "FAIL" for r in report["failed_rules"])


def test_failed_rules_match_rejection_reasons() -> None:
    """The report and the verdict cannot disagree — both derive from one pass."""
    reported = {r["rule_id"] for r in EVALUATION["evaluation_report"]["BOI"]["failed_rules"]}
    assert reported == {r["rule_id"] for r in EVALUATION["rejection_reasons"]}


# --- Binary generation ----------------------------------------------------- #


def test_excel_is_a_valid_workbook_with_verdict_fills() -> None:
    stream = build_excel({"Application ID": APPLICATION_ID}, EVALUATION["evaluation_report"])
    assert zipfile.is_zipfile(stream)

    stream.seek(0)
    book = openpyxl.load_workbook(stream)
    assert book.sheetnames[0] == "Summary"
    assert "BOI" in book.sheetnames

    sheet = book["BOI"]
    fills = {
        sheet.cell(row=r, column=1).fill.fgColor.rgb for r in range(2, sheet.max_row + 1)
    }
    assert "00D9F2E3" in fills and "00FBD9D9" in fills, "PASS green and FAIL red both present"


def test_pdf_is_a_valid_document() -> None:
    data = build_pdf({"Application ID": APPLICATION_ID}, EVALUATION["evaluation_report"]).getvalue()
    assert data.startswith(b"%PDF-")
    assert b"%%EOF" in data[-1024:]


def test_exports_never_contain_an_unmasked_pan() -> None:
    params = {"Applicant": "Rohan Sharma", "PAN (masked)": "AB******4F"}
    for stream in (build_excel(params, EVALUATION["evaluation_report"]),
                   build_pdf(params, EVALUATION["evaluation_report"])):
        assert b"ABCDE1234F" not in stream.getvalue()


# --- Endpoint -------------------------------------------------------------- #


@pytest.mark.parametrize(
    "fmt,suffix,media",
    [
        ("pdf", "pdf", "application/pdf"),
        ("excel", "xlsx",
         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    ],
)
def test_export_streams_the_document(fmt: str, suffix: str, media: str) -> None:
    _stage()
    response = client.get(
        f"/api/v1/onboarding/applications/{APPLICATION_ID}/export",
        params={"format": fmt},
        headers={"X-Tenant-ID": "tenant_beta"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(media)
    assert f'filename="flowbre-eligibility-{APPLICATION_ID}.{suffix}"' in \
        response.headers["content-disposition"]
    assert len(response.content) > 1000


def test_export_rejects_an_unknown_format() -> None:
    _stage()
    response = client.get(
        f"/api/v1/onboarding/applications/{APPLICATION_ID}/export",
        params={"format": "csv"}, headers={"X-Tenant-ID": "tenant_beta"},
    )
    assert response.status_code == 422


def test_export_404s_for_a_missing_application() -> None:
    _stage(app_row=None)
    response = client.get(
        "/api/v1/onboarding/applications/does-not-exist/export",
        headers={"X-Tenant-ID": "tenant_beta"},
    )
    assert response.status_code == 404


def test_export_409s_when_no_report_was_stored() -> None:
    """An application evaluated before the report existed cannot be exported."""
    _stage(exec_row=RuleExecutionModel(
        application_id=APPLICATION_ID, tenant_id="tenant_beta", bank_code="BOI",
        eligible=False, execution_time_ms=1.0, evaluation_report_json=None,
    ))
    response = client.get(
        f"/api/v1/onboarding/applications/{APPLICATION_ID}/export",
        headers={"X-Tenant-ID": "tenant_beta"},
    )
    assert response.status_code == 409
