"""Conformance test: the onboarding API must agree with the bank policy
spreadsheet, cell for cell.

`Bank_Eligibility_Matrix_v1.xlsx` is the contract. This module parses it into an
independent reference evaluator and replays edge-case submissions through
POST /api/v1/onboarding/evaluate/form, asserting the API's 8-bank verdict map
matches the sheet's. It is a differential test on purpose: it shares no code
with the engine, so a policy column the engine forgets to evaluate shows up as
a failure rather than passing by construction.

It exists because ten sheet columns were once declared in the policy matrix and
never evaluated (loan enquiry, rental income, HUF, agriculture, sibling
co-applicant, business ITR years, unfiled ITR, and the with/without-guarantor
split), and a hand-transcribed constant silently disagreed with its cell.

Four layers, each closing a hole the previous one leaves open:
  * policy constants    — every threshold equals its cell (catches transcription)
  * endpoint verdicts   — the API's answer over the wire (catches the route/schema)
  * per-bank verdicts   — every bank scored directly (catches the three banks the
                          form cannot select, and de-vacuizes rejecting cases)
  * column coverage     — every column is evaluated, provably inert, or an
                          explicitly documented deferral (catches silent omission)
"""

import asyncio
import re
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from unittest.mock import AsyncMock, MagicMock

import openpyxl
import pytest
from fastapi.testclient import TestClient

import app.core.redis as redis_module
from app.api.deps import get_db, get_redis
from app.api.schemas.onboarding import OnboardingFormRequest
from app.constants.form_mappings import EXISTING_BANK_TO_BANK_CODE
from app.constants.limits import MIN_SELF_EMPLOYED_COMBINED_ITR as COMBINED_ITR_FLOOR
from app.main import app
from app.services.bre_engine import BANK_MATRIX_RULES, RENTAL_CLASS_TO_FLAG, bre_engine_service

ZEN_RULES = Path(__file__).resolve().parents[2] / "app" / "zen_rules"
# The matrix is split by entity type. Policy is read by COLUMN HEADER, not
# position, because the two files carry different column orders and widths.
INDIVIDUAL_SHEET = ZEN_RULES / "bank_Individual_Eligibility_Matrix.xlsx"
COMPANY_SHEET = ZEN_RULES / "bank_Company_Organization_Eligibility_Matrix.xlsx"
# 24 columns (bureau floors, write-off caps, DPD) necessarily appear in BOTH
# files — a CIBIL floor binds a company as much as an individual. That is the
# split's one hazard, guarded by test_split_matrices_agree_on_shared_columns.
SHEET_PATH = INDIVIDUAL_SHEET
TENANT = "tenant_beta"
TODAY = date.today()

SHEET_NAME_TO_CODE = {
    "BOI": "BOI",
    "Indian Bank": "INDIAN_BANK",
    "IOB": "IOB",
    "BOB": "BOB",
    "BOM": "BOM",
    "HDFC": "HDFC",
    "AXIS": "AXIS",
    "Kotak": "KOTAK",
}
# Only these five are reachable as the *selected* bank from the form's
# "Existing Current/Savings Bank A/c With" options. HDFC / AXIS / Kotak are
# covered by test_every_bank_verdict_matches_spreadsheet, which scores every
# bank directly — the eligibility map alone cannot verify them, because it is
# ANDed with the selected bank's verdict.
# Bank code -> the `existingAccountBank` option that selects it. Derived from
# the mapping table rather than restated, so a bank added to one and not the
# other cannot silently drop out of the endpoint conformance loop below.
FORM_SELECTABLE = {code.value: option.value
                   for option, code in EXISTING_BANK_TO_BANK_CODE.items()}


# --------------------------------------------------------------------------- #
# 1. Parse the spreadsheet into per-bank policy dicts
# --------------------------------------------------------------------------- #

def _flag(v: Any) -> bool:
    return str(v).strip().lower() == "true"


def _threshold(v: Any) -> float:
    """'>= 701' -> 701, '< 5000' -> 5000, '<= 0' -> 0."""
    m = re.search(r"(-?\d+(?:\.\d+)?)", str(v))
    return float(m.group(1)) if m else 0.0


def _max_acceptable(v: Any) -> int:
    """'<= 0' -> 0 (0 is acceptable); '< 90' -> 89 (90 rejects)."""
    text = str(v).strip()
    n = int(_threshold(text))
    return n if text.startswith("<=") else n - 1


def _cell(row: Dict[str, Any], *candidates: str) -> Any:
    """Read a policy cell by any of its accepted header spellings.

    The entity matrices are hand-curated and columns get renamed to scope them
    ("Business Proof" -> "Self-employed-Business Proof"; the Company sheet drops
    the "SE " prefix entirely). Matching a list of spellings — whitespace- and
    case-insensitively — keeps a rename from breaking collection, so only a
    genuinely NEW column requires a code change.
    """
    normalized = {str(k).strip().lower(): v for k, v in row.items()}
    for name in candidates:
        if (key := name.strip().lower()) in normalized:
            return normalized[key]
    raise KeyError(f"none of {candidates!r} present; sheet has {sorted(row)}")


def _read(path: Path) -> Dict[str, Dict[str, Any]]:
    """{bank_code: {column_header: cell}} for one entity-scoped matrix."""
    rows = list(openpyxl.load_workbook(str(path), data_only=True)["decision table"].iter_rows(values_only=True))
    header = [str(h) for h in rows[0]]
    return {
        SHEET_NAME_TO_CODE[row[0]]: dict(zip(header, row))
        for row in rows[1:]
    }


def load_policies() -> Dict[str, Dict[str, Any]]:
    """Merge both entity-scoped matrices into the per-bank policy the engine
    must satisfy. Individual-only and Company-only columns come from their own
    file; shared columns resolve identically from either."""
    individual, company = _read(INDIVIDUAL_SHEET), _read(COMPANY_SHEET)
    policies: Dict[str, Dict[str, Any]] = {}

    for code in individual:
        ind, com = individual[code], company[code]
        policies[code] = {
            "min_cibil": _threshold(_cell(ind, "CIBIL Score")),
            "allow_write_off": {
                "PL": _flag(_cell(ind, "PL Write off")),
                "HL": _flag(_cell(ind, "Home Loan Write off")),
                "CONSUMER": _flag(_cell(ind, "Consumer Loan  Write off")),
                "AGRI": _flag(_cell(ind, "Agri Loan  Write off")),
                "MSME": _flag(_cell(ind, "MSME Loan  Write off")),
                "AUTO": _flag(_cell(ind, "Auto Loan  Write off")),
                "CC": _flag(_cell(ind, "Credit Card Write Off History", "Credit Card Write Off")),
            },
            "max_cc_write_off": _threshold(_cell(ind, "Credit Card Write Off Amount history", "Credit Card Write Off Amount")),
            "max_dpd": _max_acceptable(_cell(ind, "DPD")),
            "allow_loan_enquiry": _flag(_cell(ind, "Loan enquiry")),
            "allow_currently_outstanding": _flag(_cell(ind, "Currently Outstanding")),
            "min_age": int(_cell(ind, "Min Age")),
            "max_age_emi_salaried": _threshold(_cell(ind, "Age at Last EMI-Salaried")),
            "max_age_emi_self_employed": _threshold(_cell(ind, "Age at Last EMI-Self Employed")),
            "allow_existing_car_loan": _flag(_cell(ind, "Existing Car Loan")),
            "requires_existing_account": _flag(_cell(ind, "Existing A/C Holder")),
            "allow_separate_both_rented": _flag(_cell(ind, "Resi-Office-Separate-Both Rented-selfemployee", "Resi-Office-Separate-Both Rented")),
            "allow_without_guarantor": _flag(_cell(ind, "Without a Guarantor")),
            "allow_with_guarantor": _flag(_cell(ind, "With a Guarantor")),
            "allow_nri": _flag(_cell(ind, "NRI/PIO")),
            "min_nri_stay_years": _threshold(_cell(ind, "Minimium Stay Period for NRI")),
            "allow_agriculture": _flag(_cell(ind, "Agriculture-selfemployee", "Agriculture")),
            "min_total_experience_years": _threshold(_cell(ind, "Minimum work experience")),
            "min_current_company_years": _threshold(_cell(ind, "Current Company Experience (Years)")),
            "allow_cash_salary": _flag(_cell(ind, "Salary payment mode-Cash")),
            "allow_no_income_proof": _flag(_cell(ind, "No Income Proof")),
            "allow_rental": {
                "NO_ITR_NOT_IN_BANK": _flag(_cell(ind, "Rental Income - With Agreement - Not Filed ITR - Not Reflecting in Bank")),
                "NO_ITR_IN_BANK": _flag(_cell(ind, "Rental Income - With Agreement - Not Filed ITR - Reflecting in Bank")),
                "ITR_NOT_IN_BANK": _flag(_cell(ind, "Rental Income - With Agreement - Filed ITR - Not Reflecting in Bank")),
            },
            "min_salary": _threshold(_cell(ind, "Minimum Salary")),
            # Income rules are read from the COMPANY sheet: they are the rules a
            # Company is scored on, and the split must keep them authoritative
            # there. They are identical in the Individual sheet (shared columns).
            "se_min_current_itr": _threshold(_cell(com, "SE Current ITR", "Current ITR")),
            "se_combined_itr": "Current + Prev" in str(_cell(com, "SE Previous ITR", "Previous ITR")),
            "se_min_prev_itr": _threshold(_cell(com, "SE Previous ITR", "Previous ITR")),
            "allow_itr_not_filed": _flag(_cell(ind, "Self Employed-ITR Not Filed", "ITR Not Filed")),
            "min_business_itr_years": _threshold(_cell(com, "Business ITR Years", "Self-employed-Business ITR Years")),
            "allow_huf": _flag(_cell(ind, "HUF")),
            "form16_years_required": _threshold(_cell(ind, "salaried-Form 16 Years", "Form 16 Years")),
            "coapp": {
                "AGE_Brother": _flag(_cell(ind, "Co-Applicant Age-Brother")),
                "AGE_Sister": _flag(_cell(ind, "Co-Applicant Age-Sister")),
                "INC_Brother": _flag(_cell(ind, "Co-Applicant Income-Brother")),
                "INC_Father": _flag(_cell(ind, "Co-Applicant Income-Father")),
                "INC_Mother": _flag(_cell(ind, "Co-Applicant Income-Mother")),
                "INC_Sister": _flag(_cell(ind, "Co-Applicant Income-Sister")),
            },
        }
    return policies


POLICIES = load_policies()


# --------------------------------------------------------------------------- #
# 2. Reference evaluator — the spreadsheet's verdict for one applicant
# --------------------------------------------------------------------------- #
#
# Tie-break note: cols 21/22/23 are mutually inconsistent for BOB and HDFC
# (col21 permits the both-rented configuration while col22/23 refuse it either
# way). The guarantor-specific columns 22/23 are treated as authoritative
# because they address the guarantor question directly; col21 is read as a
# descriptive flag. Under that reading only BOM waives the guarantor, and
# IOB/BOB decline the configuration outright.


def sheet_rejects(bank: str, f: Dict[str, Any]) -> List[str]:
    p = POLICIES[bank]
    out: List[str] = []

    # Demographics
    if f["age"] < p["min_age"]:
        out.append("min_age")
    if f["is_nri"]:
        if not p["allow_nri"]:
            out.append("nri_not_allowed")
        elif f["nri_stay_years"] < p["min_nri_stay_years"]:
            out.append("nri_stay")

    # Bureau
    if f["cibil"] < p["min_cibil"]:
        out.append("cibil")
    if f["dpd"] > p["max_dpd"]:
        out.append("dpd")
    if f["currently_outstanding"] > 0 and not p["allow_currently_outstanding"]:
        out.append("currently_outstanding")
    if f["loan_enquiry"] and not p["allow_loan_enquiry"]:
        out.append("loan_enquiry")
    if f["write_off_amount"] > 0:
        wtype = f["write_off_type"]
        if wtype is None:
            out.append("write_off_unclassified")
        elif not p["allow_write_off"][wtype]:
            out.append(f"write_off_{wtype}")
        elif wtype == "CC" and f["write_off_amount"] >= p["max_cc_write_off"]:
            out.append("write_off_cc_cap")

    # Entity / business classification
    if f["is_huf"] and not p["allow_huf"]:
        out.append("huf")
    if f["is_agriculture"] and not p["allow_agriculture"]:
        out.append("agriculture")

    # Employment & income
    if f["occupation"] == "Salaried":
        if f["salary"] < p["min_salary"]:
            out.append("min_salary")
        if f["cash_salary"] and not p["allow_cash_salary"]:
            out.append("cash_salary")
        if f["work_exp_years"] < p["min_total_experience_years"]:
            out.append("work_experience")
        if f["current_company_years"] < p["min_current_company_years"]:
            out.append("company_tenure")
        if f["no_income_proof"]:
            if not p["allow_no_income_proof"]:
                out.append("no_income_proof")
        elif f["form_16_years"] < p["form16_years_required"]:
            out.append("form16")
        if f["age_at_last_emi"] > p["max_age_emi_salaried"]:
            out.append("age_emi_salaried")
    else:
        if f["business_itr_years"] < p["min_business_itr_years"]:
            out.append("business_itr_years")
        if not f["itr_filed"]:
            if not p["allow_itr_not_filed"]:
                out.append("itr_not_filed")
        else:
            if p["se_combined_itr"]:
                # The two-year total is the whole income test at these banks;
                # the per-year floors do not also apply.
                if f["current_itr"] + f["previous_itr"] < COMBINED_ITR_FLOOR:
                    out.append("se_combined_itr")
            else:
                if f["current_itr"] < p["se_min_current_itr"]:
                    out.append("se_current_itr")
                if f["previous_itr"] < p["se_min_prev_itr"]:
                    out.append("se_prev_itr")
        if not f["business_proof"]:
            out.append("business_proof")
        if f["age_at_last_emi"] > p["max_age_emi_self_employed"]:
            out.append("age_emi_self_employed")

    # Secondary rental income
    if f["rental"] and not p["allow_rental"][f["rental"]]:
        out.append(f"rental_{f['rental']}")

    # Residence / office tenure & guarantor
    # Separate premises, both rented (col 21) — distinct from the guarantor
    # question, which governs an office run out of a rented residence.
    if f["separate_both_rented"] and not p["allow_separate_both_rented"]:
        out.append("separate_both_rented")

    if f["resi_cum_office_rented"]:
        permitted = p["allow_with_guarantor"] if f["guarantor"] else p["allow_without_guarantor"]
        if not permitted:
            out.append("both_rented_guarantor")

    # Existing banking relationship — every bank but IOB lends only to its own
    # existing account holders (col 17).
    if p["requires_existing_account"] and f["existing_account_bank"] != bank:
        out.append("existing_account")

    if f["car_loan"] and not p["allow_existing_car_loan"]:
        out.append("existing_car_loan")

    # Co-applicant relationships
    for key in f["coapp"]:
        if not p["coapp"][key]:
            out.append(f"coapp_{key}")

    return out


# --------------------------------------------------------------------------- #
# 3. Case construction: semantic facts -> (form payload, derived facts)
# --------------------------------------------------------------------------- #

RENTAL_FORM_VALUE = {
    "NO_ITR_NOT_IN_BANK": "Rental Income-with Agreement -Not filed ITR-Not reflecting in Bank",
    "ITR_NOT_IN_BANK": "Rental Income-with Agreement filed ITR- Not reflecting in Bank",
    "NO_ITR_IN_BANK": "Rental Income-with Agreement -Not filed ITR-reflecting in Bank",
}
WRITE_OFF_FORM_FLAG = {
    "PL": "bureauFlagPL", "HL": "bureauFlagHome", "CONSUMER": "bureauFlagConsumer",
    "AGRI": "bureauFlagAgri", "MSME": "bureauFlagMSME", "AUTO": "bureauFlagAuto",
    "CC": "bureauFlagCC",
}

DEFAULTS: Dict[str, Any] = {
    "entity": "Individual", "occupation": "Salaried", "selected": "BOI",
    "age": 34, "is_nri": False, "nri_months": 0,
    "cibil": 800, "dpd": 0, "currently_outstanding": 0.0, "loan_enquiry": False,
    "write_off_type": None, "write_off_amount": 0.0,
    "age_at_last_emi": 55, "car_loan": False,
    "salary_band": "gt25000", "cash_salary": False, "tenure_band": "2y+", "form16_years": 2,
    "prev_joining_years_ago": None, "no_income_proof": False, "rental": None,
    "current_itr": 500000.0, "previous_itr": 350000.0, "itr_filed": True,
    "business_years_ago": 10, "business_itr_years": 5, "business_proof": True, "business_entity": "Propreitorship",
    "residence": "Owned House", "office": None, "office_status": None, "guarantor": None,
    "coapp_age": "None", "coapp_income": "None",
}

TENURE_MONTHS = {"0-6m": 0, "6m-1y": 6, "1y-2y": 12, "2y+": 24}


def derive(c: Dict[str, Any]) -> Dict[str, Any]:
    """The numeric inputs the engine is expected to see for this case."""
    months = TENURE_MONTHS[c["tenure_band"]]
    work_exp = c["prev_joining_years_ago"] if c["prev_joining_years_ago"] else months // 12
    is_huf = c["entity"] == "HUF" or c["business_entity"] == "HUF"
    coapp = []
    if c["coapp_age"] != "None":
        coapp.append(f"AGE_{c['coapp_age']}")
    if c["coapp_income"] != "None":
        coapp.append(f"INC_{c['coapp_income']}")
    return {
        "age": c["age"],
        "is_nri": c["is_nri"],
        "nri_stay_years": c["nri_months"] / 12.0,
        "cibil": c["cibil"],
        "dpd": c["dpd"],
        "currently_outstanding": c["currently_outstanding"],
        "loan_enquiry": bool(c["loan_enquiry"]),
        "write_off_type": c["write_off_type"],
        "write_off_amount": c["write_off_amount"],
        "is_huf": is_huf,
        "is_agriculture": c["business_entity"] == "Agriculture",
        "occupation": c["occupation"],
        "salary": 50000.0 if c["salary_band"] == "gt25000" else 20000.0,
        "cash_salary": c["cash_salary"],
        "work_exp_years": work_exp,
        "current_company_years": months / 12.0,
        "no_income_proof": c["no_income_proof"],
        "form_16_years": 0 if c["no_income_proof"] else c["form16_years"],
        "age_at_last_emi": c["age_at_last_emi"],
        "business_years": c["business_years_ago"],
        # HUF collects no explicit filing count; the engine falls back to
        # business age there, so the reference must too.
        "business_itr_years": (
            c["business_years_ago"] if c["entity"] == "HUF" else c["business_itr_years"]
        ),
        "itr_filed": c["itr_filed"],
        "current_itr": c["current_itr"],
        "previous_itr": c["previous_itr"],
        "business_proof": c["business_proof"],
        "rental": c["rental"],
        # The office runs out of a rented residence — the only guarantor trigger.
        "resi_cum_office_rented": c["residence"] == "Rented House" and c["office"] == "Same",
        "separate_both_rented": (c["residence"] == "Rented House" and c["office"] == "Separate"
                                 and c["office_status"] == "Rented"),
        "guarantor": c["guarantor"] == "With a Gaurantor",
        "car_loan": c["car_loan"],
        "existing_account_bank": c["selected"],
        "coapp": coapp,
    }


def build_form(c: Dict[str, Any]) -> Dict[str, Any]:
    dob = TODAY.replace(year=TODAY.year - c["age"])
    est = TODAY.replace(year=TODAY.year - c["business_years_ago"])

    banking: Dict[str, Any] = {
        "existingAccountBank": FORM_SELECTABLE[c["selected"]],
        "existingCarLoanBank": "BOB" if c["car_loan"] else "None",
        "loanType": "Auto Loan",
        "bureauCibilScore": c["cibil"],
        "bureauDpd": c["dpd"],
        "bureauLoanEnquiry": bool(c["loan_enquiry"]),
        "bureauCurrentlyOutstanding": c["currently_outstanding"],
        "bureauAgeAtLastEMI": c["age_at_last_emi"],
    }
    if c["write_off_type"]:
        banking[WRITE_OFF_FORM_FLAG[c["write_off_type"]]] = True
        banking["bureauWriteOffAmount"] = c["write_off_amount"]

    if c["entity"] == "HUF":
        identity = {
            "entityType": "HUF", "applicantName": "Sharma HUF", "hufName": "Sharma HUF",
            "hufPan": "AAAHS1234F", "kartaName": "Rakesh Sharma", "kartaPan": "ABCDE1234F",
        }
        occupation = {
            "profileType": "HUF", "officeAddressType": "Same",
            "businessEstablishmentDate": est.isoformat(),
            "itrFilingStatus": "Self employed ITR Filled" if c["itr_filed"] else "ITR Not Filed",
            "businessProof": "GSTIN: 29AAAAA0000A1Z5" if c["business_proof"] else None,
        }
        if c["itr_filed"]:
            occupation["currentITRAmount"] = c["current_itr"]
            occupation["prevITRAmount"] = c["previous_itr"]
        if c["rental"]:
            occupation["rentalIncomeTypeSelfEmployed"] = RENTAL_FORM_VALUE[c["rental"]]
    else:
        identity = {
            "entityType": "Individual", "applicantName": "Rohan Sharma",
            "dob": dob.isoformat(), "pan": "ABCDE1234F", "phone": "9876543210",
            "email": "rohan.sharma@example.com",
            "citizenshipStatus": "NRI/PIO" if c["is_nri"] else "Resident Indian",
        }
        if c["is_nri"]:
            identity["nriStayPeriod"] = c["nri_months"]

        if c["occupation"] == "Salaried":
            occupation = {
                "profileType": "Salaried", "tenureBand": c["tenure_band"],
                "grossSalary": 50000.0 if c["salary_band"] == "gt25000" else 20000.0,
                "salaryMode": "Salary payment mode-Cash" if c["cash_salary"]
                              else "Salary payment mode- Bank Credit",
                "form16Status": "No Income Proof" if c["no_income_proof"] else "Form 16",
            }
            if not c["no_income_proof"]:
                occupation["form16Years"] = c["form16_years"]
            if c["prev_joining_years_ago"]:
                prev = TODAY.replace(year=TODAY.year - c["prev_joining_years_ago"])
                occupation["prevCompanyName"] = "Prior Employer Pvt Ltd"
                occupation["prevCompanyJoining"] = prev.isoformat()
            if c["rental"]:
                occupation["rentalIncomeTypeSalaried"] = RENTAL_FORM_VALUE[c["rental"]]
        else:
            occupation = {
                "profileType": "Self-Employed",
                "businessEntityType": c["business_entity"],
                "businessEstablishmentDate": est.isoformat(),
                "currentITRAmount": c["current_itr"], "prevITRAmount": c["previous_itr"],
                "businessItrAmount": c["business_itr_years"],
                "businessProof": "GSTIN: 29AAAAA0000A1Z5" if c["business_proof"] else None,
            }
            if c["office"]:
                occupation["officeAddressType"] = c["office"]
                if c["office"] == "Separate":
                    occupation["officeAddress"] = "Indiranagar, Bengaluru"
                    occupation["officePremisesStatus"] = c["office_status"]
            if c["guarantor"]:
                occupation["guarantorStatus"] = c["guarantor"]
            if c["rental"]:
                occupation["rentalIncomeTypeSelfEmployed"] = RENTAL_FORM_VALUE[c["rental"]]

    payload: Dict[str, Any] = {
        "identity": identity,
        "address": {"pincode": "560001", "residentDetails": c["residence"]},
        "occupation": occupation,
        "banking": banking,
        "coApplicant": {"coAppAgeRelation": c["coapp_age"], "coAppIncomeRelation": c["coapp_income"]},
    }
    if c["coapp_income"] != "None":
        payload["coApplicant"].update({
            "coApplicantName": "Anil Sharma", "coApplicantDob": "1965-03-10",
            "coApplicantOccupation": "Salaried",
        })
    return payload


# --------------------------------------------------------------------------- #
# 4. Test cases — one clean baseline plus a targeted mutation per policy column
# --------------------------------------------------------------------------- #

SELF_EMPLOYED = {"occupation": "Self-Employed", "office": "Same"}


def cases() -> List[Tuple[str, Dict[str, Any]]]:
    out: List[Tuple[str, Dict[str, Any]]] = []

    def add(name: str, **over: Any) -> None:
        out.append((name, {**DEFAULTS, **over}))

    for bank in FORM_SELECTABLE:
        b = {"selected": bank}
        floor = int(POLICIES[bank]["min_cibil"])
        add(f"{bank}/clean-salaried", **b)
        add(f"{bank}/clean-self-employed", **b, **SELF_EMPLOYED)
        add(f"{bank}/cibil-at-floor", **b, cibil=floor)
        add(f"{bank}/cibil-below-floor", **b, cibil=floor - 1)
        add(f"{bank}/dpd-1", **b, dpd=1)
        add(f"{bank}/dpd-89", **b, dpd=89)
        add(f"{bank}/dpd-90", **b, dpd=90)
        add(f"{bank}/cc-write-off-4999", **b, write_off_type="CC", write_off_amount=4999.0)
        add(f"{bank}/cc-write-off-5000", **b, write_off_type="CC", write_off_amount=5000.0)
        add(f"{bank}/cc-write-off-9999", **b, write_off_type="CC", write_off_amount=9999.0)
        add(f"{bank}/pl-write-off", **b, write_off_type="PL", write_off_amount=1000.0)
        add(f"{bank}/auto-write-off", **b, write_off_type="AUTO", write_off_amount=1000.0)
        add(f"{bank}/currently-outstanding", **b, currently_outstanding=15000.0)
        add(f"{bank}/loan-enquiry", **b, loan_enquiry=True)
        add(f"{bank}/age-20", **b, age=20)
        add(f"{bank}/emi-age-61", **b, age_at_last_emi=61)
        add(f"{bank}/emi-age-71", **b, age_at_last_emi=71)
        add(f"{bank}/emi-age-66-self-employed", **b, **SELF_EMPLOYED, age_at_last_emi=66)
        add(f"{bank}/cash-salary", **b, cash_salary=True)
        add(f"{bank}/salary-below-floor", **b, salary_band="lt25000")
        add(f"{bank}/no-income-proof", **b, no_income_proof=True)
        add(f"{bank}/form16-1-year", **b, form16_years=1)
        add(f"{bank}/tenure-1y", **b, tenure_band="1y-2y", prev_joining_years_ago=4)
        add(f"{bank}/tenure-6m", **b, tenure_band="6m-1y", prev_joining_years_ago=4)
        add(f"{bank}/car-loan", **b, car_loan=True)
        add(f"{bank}/nri-24-months", **b, is_nri=True, nri_months=24)
        add(f"{bank}/nri-12-months", **b, is_nri=True, nri_months=12)
        for rental in RENTAL_FORM_VALUE:
            add(f"{bank}/rental-{rental}", **b, rental=rental)
        add(f"{bank}/se-current-itr-low", **b, **SELF_EMPLOYED, current_itr=90000.0)
        add(f"{bank}/se-prev-itr-low", **b, **SELF_EMPLOYED, previous_itr=90000.0)
        add(f"{bank}/se-itr-years-1", **b, **SELF_EMPLOYED, business_itr_years=1)
        add(f"{bank}/se-no-business-proof", **b, **SELF_EMPLOYED, business_proof=False)
        add(f"{bank}/se-agriculture", **b, **SELF_EMPLOYED, business_entity="Agriculture")
        add(f"{bank}/huf-entity", **b, entity="HUF")
        add(f"{bank}/huf-itr-not-filed", **b, entity="HUF", itr_filed=False)
        add(f"{bank}/resi-cum-office-rented-with-guarantor", **b, **SELF_EMPLOYED,
            residence="Rented House", guarantor="With a Gaurantor")
        add(f"{bank}/resi-cum-office-rented-without-guarantor", **b, **SELF_EMPLOYED,
            residence="Rented House", guarantor="Without a Gaurantor")
        # A separately addressed office no longer prompts for a guarantor.
        add(f"{bank}/separate-office-both-rented", **b, occupation="Self-Employed",
            residence="Rented House", office="Separate", office_status="Rented")
        add(f"{bank}/coapp-age-brother", **b, coapp_age="Brother")
        add(f"{bank}/coapp-income-sister", **b, coapp_income="Sister")
        add(f"{bank}/coapp-income-father", **b, coapp_income="Father")

    return out


# --------------------------------------------------------------------------- #
# 5. Test harness
# --------------------------------------------------------------------------- #


async def _mock_db():
    session = MagicMock()
    fut: Any = asyncio.Future()
    fut.set_result(MagicMock())
    session.execute.return_value = fut
    session.flush.return_value = fut
    session.commit.return_value = fut
    session.rollback.return_value = fut
    yield session


mock_redis = AsyncMock()
mock_redis.incr.return_value = 1
mock_redis.expire.return_value = True
mock_redis.get.return_value = None
mock_redis.setex.return_value = True
redis_module.redis_client = mock_redis


async def _mock_redis_dep():
    return mock_redis


app.dependency_overrides[get_db] = _mock_db
app.dependency_overrides[get_redis] = _mock_redis_dep
client = TestClient(app)

CASES = cases()

# Sheet column -> the BANK_MATRIX_RULES key that must carry its value.
CONSTANT_CORRESPONDENCE = [
    ("min_cibil", "min_cibil"),
    ("max_cc_write_off", "max_cc_write_off_amount"),
    ("max_dpd", "max_dpd"),
    ("allow_loan_enquiry", "allow_loan_enquiry"),
    ("requires_existing_account", "requires_existing_account"),
    ("min_age", "min_age"),
    ("max_age_emi_salaried", "max_age_emi_salaried"),
    ("max_age_emi_self_employed", "max_age_emi_self_employed"),
    ("allow_without_guarantor", "allow_without_guarantor"),
    ("allow_with_guarantor", "allow_with_guarantor"),
    ("allow_nri", "allow_nri"),
    ("min_nri_stay_years", "min_nri_stay_years"),
    ("allow_agriculture", "allow_agriculture"),
    ("min_total_experience_years", "min_total_experience_years"),
    ("min_current_company_years", "min_current_company_tenure_years"),
    ("allow_no_income_proof", "allow_no_income_proof"),
    ("min_salary", "min_salary"),
    ("se_min_current_itr", "se_min_current_itr"),
    ("allow_itr_not_filed", "allow_itr_not_filed"),
    ("min_business_itr_years", "min_business_itr_years"),
    ("allow_huf", "allow_huf"),
    ("form16_years_required", "form16_years_required"),
]


@pytest.mark.parametrize("bank", sorted(SHEET_NAME_TO_CODE.values()))
def test_policy_constants_match_spreadsheet(bank: str) -> None:
    """Every threshold and flag in the code matrix equals its spreadsheet cell."""
    sheet, code_policy = POLICIES[bank], BANK_MATRIX_RULES[bank]

    for sheet_key, code_key in CONSTANT_CORRESPONDENCE:
        assert float(sheet[sheet_key]) == float(code_policy[code_key]), (
            f"{bank}.{code_key}: sheet={sheet[sheet_key]} code={code_policy[code_key]}"
        )

    for rental_class, allowed in sheet["allow_rental"].items():
        code_key = RENTAL_CLASS_TO_FLAG[rental_class]
        assert code_policy[code_key] == allowed, f"{bank}.{code_key}"

    for write_off_type, allowed in sheet["allow_write_off"].items():
        code_key = {"PL": "allow_pl_write_off", "HL": "allow_hl_write_off",
                    "CONSUMER": "allow_consumer_write_off", "AGRI": "allow_agri_write_off",
                    "MSME": "allow_msme_write_off", "AUTO": "allow_auto_write_off",
                    "CC": "allow_cc_write_off"}[write_off_type]
        assert code_policy[code_key] == allowed, f"{bank}.{code_key}"

    # Cols 56-58/61 collapse to one flag only while they agree with each other.
    sibling = {sheet["coapp"][k] for k in ("AGE_Brother", "AGE_Sister", "INC_Brother", "INC_Sister")}
    assert len(sibling) == 1, f"{bank}: sheet disagrees across the sibling co-applicant columns"
    assert code_policy["allow_sibling_coapplicant"] == sibling.pop(), (
        f"{bank}.allow_sibling_coapplicant"
    )


@pytest.mark.parametrize("case", [c for _, c in CASES], ids=[n for n, _ in CASES])
def test_endpoint_verdict_matches_spreadsheet(case: Dict[str, Any]) -> None:
    """The API's verdict — and every bank in its eligibility map — matches the sheet."""
    facts = derive(case)
    response = client.post(
        "/api/v1/onboarding/evaluate/form",
        json=build_form(case),
        headers={"X-Tenant-ID": TENANT},
    )
    assert response.status_code == 200, response.text
    body = response.json()

    selected = case["selected"]
    expected_overall = not sheet_rejects(selected, facts)
    assert body["overall_eligible"] == expected_overall, (
        f"selected={selected} sheet_says={sheet_rejects(selected, facts) or ['<eligible>']} "
        f"engine_says={[r['rule_id'] for r in body['rejection_reasons']] or ['<eligible>']}"
    )

    # Each entry is that bank's OWN verdict, independent of the selected bank.
    # These assertions used to AND in `expected_overall`, mirroring an engine
    # bug: once the selected bank rejected, every expectation collapsed to
    # False and the loop asserted nothing about the other seven banks.
    for code in POLICIES:
        expected = not sheet_rejects(code, facts)
        assert body["bank_eligibility"][code] == expected, (
            f"bank_eligibility[{code}] sheet_says={sheet_rejects(code, facts) or ['<eligible>']} "
            f"engine_says={body['bank_eligibility'][code]}; selected={selected}"
        )
        report = body["evaluation_report"][code]
        assert report["is_eligible"] == expected
        # A bank cannot be turned down without naming a rule (the old AND
        # produced is_eligible=False alongside an empty failure list).
        assert bool(report["failed_rules"]) != expected, (
            f"{code}: is_eligible={report['is_eligible']} but failed_rules="
            f"{[r['rule_id'] for r in report['failed_rules']]}"
        )


# --------------------------------------------------------------------------- #
# 6. Independent per-bank verification
# --------------------------------------------------------------------------- #
#
# The endpoint's bank_eligibility map is each bank's verdict ANDed with the
# selected bank's, so once the selected bank rejects, every entry is False and
# the map assertions above can no longer distinguish a correct policy from a
# broken one. These tests evaluate each bank as the selected bank directly —
# un-ANDed — which is also the only way HDFC / AXIS / Kotak get covered, since
# the form's bank list cannot select them.

SCENARIOS = [(name.split("/", 1)[1], case) for name, case in CASES if case["selected"] == "BOI"]


@pytest.mark.parametrize("case", [c for _, c in SCENARIOS], ids=[n for n, _ in SCENARIOS])
def test_every_bank_verdict_matches_spreadsheet(case: Dict[str, Any]) -> None:
    """Each of the 8 banks, scored independently on the same applicant."""
    facts = derive(case)
    engine_payload = OnboardingFormRequest.model_validate(build_form(case)).to_engine_payload()

    for bank in POLICIES:
        payload = {**engine_payload, "selected_bank": bank}
        verdict = asyncio.run(bre_engine_service.evaluate_application(payload, tenant_id="default"))
        expected = not sheet_rejects(bank, facts)
        assert verdict["overall_eligible"] == expected, (
            f"{bank}: sheet_says={sheet_rejects(bank, facts) or ['<eligible>']} "
            f"engine_says={[r['rule_id'] for r in verdict['rejection_reasons']] or ['<eligible>']}"
        )


# --------------------------------------------------------------------------- #
# 7. Column coverage — no policy column may be silently ignored
# --------------------------------------------------------------------------- #

METADATA_COLUMNS = {"Bank Name", "Description"}

# Columns with a rejection path in the engine, conformance-tested above.
EVALUATED_COLUMNS = {
    "CIBIL Score",
    "PL Write off", "Home Loan Write off", "Consumer Loan  Write off",
    "Agri Loan  Write off", "MSME Loan  Write off", "Auto Loan  Write off",
    "Credit Card Write Off History", "Credit Card Write Off history",
    "Credit Card Write Off Amount history",
    "DPD", "Loan enquiry", "Currently Outstanding",
    "Min Age", "Age at Last EMI-Salaried", "Age at Last EMI-Self Employed",
    "Existing Car Loan", "Existing A/C Holder",
    "Without a Guarantor", "With a Guarantor",
    "Resi-Office-Separate-Both Rented-selfemployee",
    "NRI/PIO", "Minimium Stay Period for NRI",
    "Agriculture-selfemployee",
    "Minimum work experience", "Current Company Experience (Years)",
    "Salary payment mode-Cash", "No Income Proof",
    "Rental Income - With Agreement - Not Filed ITR - Not Reflecting in Bank",
    "Rental Income - With Agreement - Not Filed ITR - Reflecting in Bank",
    "Rental Income - With Agreement - Filed ITR - Not Reflecting in Bank",
    "Minimum Salary", "SE Current ITR", "SE Previous ITR",
    "Current ITR", "Previous ITR",
    "Self Employed-ITR Not Filed", "ITR Not Filed",
    "Business ITR Years", "Business Proof",
    "Self-employed-Business ITR Years", "Self-employed-Business Proof",
    "HUF", "salaried-Form 16 Years",
    "Co-Applicant Age-Brother", "Co-Applicant Age-Sister",
    "Co-Applicant Income-Brother", "Co-Applicant Income-Father",
    "Co-Applicant Income-Mother", "Co-Applicant Income-Sister",
}

# Permissive for every bank, so they cannot change any verdict. Deliberately
# unimplemented — but only while they stay uniform, which this module enforces.
INERT_COLUMNS = {
    "Rented House-Salaried", "Resi-Cum-Office-Owned-selfemployee", "Unmarried",
    "salaried-Employment-Firm", "salaried-Employment-Pvt Ltd",
    "salaried-Employment-Public Ltd", "salaried-Employment-Govt",
    "salaried-Employment-PSU", "Salary payment mode- Bank Credit",
    "Self Employed", "Self Employed ITR Filed", "ITR Filed",
    "Self Employed-Propreitorship",
    "Parternship Firm", "Private Limited", "Public Limited", "EMI / Income Ratio",
}

# Columns that discriminate between banks but are NOT enforced, each pending a
# policy decision rather than an implementation.
DEFERRED_COLUMNS: dict[str, str] = {}

# The curated matrices renamed some shared columns per entity. These pairs are
# the SAME policy column under two names, so the drift guard must still compare
# them; matching on name alone would silently stop checking them.
SHARED_COLUMN_ALIASES = {
    "Credit Card Write Off History": "Credit Card Write Off history",
    "Self-employed-Business ITR Years": "Business ITR Years",
    "Self-employed-Business Proof": "Business Proof",
    "SE Current ITR": "Current ITR",
    "SE Previous ITR": " Previous ITR",
    "Self Employed ITR Filed": "ITR Filed",
}
# Deliberately entity-specific: HDFC/AXIS/Kotak decline an unfiled ITR from an
# individual but accept one from a company. Cross-checking these would assert
# the split away.
ENTITY_DIVERGENT_COLUMNS = {"Self Employed-ITR Not Filed": "ITR Not Filed"}


def _headers(path: Path) -> List[str]:
    rows = openpyxl.load_workbook(str(path), data_only=True)["decision table"].iter_rows(values_only=True)
    return [str(h) for h in next(iter(rows))]


def test_column_classification_covers_both_matrices() -> None:
    """Every column across both entity matrices is evaluated, provably inert,
    or explicitly deferred — nothing is silently ignored by the engine."""
    norm = lambda s: s.strip().lower()  # noqa: E731
    present = {norm(h) for h in _headers(INDIVIDUAL_SHEET) + _headers(COMPANY_SHEET)}
    classified = {
        norm(c) for c in
        METADATA_COLUMNS | EVALUATED_COLUMNS | INERT_COLUMNS | set(DEFERRED_COLUMNS)
    }

    assert present - classified == set(), (
        f"unclassified columns: {sorted(present - classified)} — "
        "the matrices changed shape; classify them before shipping"
    )
    assert classified - present == set(), (
        f"classified but missing from both matrices: {sorted(classified - present)} — "
        "a column was dropped by the split"
    )


def test_split_matrices_agree_on_shared_columns() -> None:
    """Columns carried by BOTH matrices must hold identical values. Two copies
    of a CIBIL floor is the split's one real hazard; this is the guard."""
    individual, company = _read(INDIVIDUAL_SHEET), _read(COMPANY_SHEET)
    ind_headers, com_headers = _headers(INDIVIDUAL_SHEET), _headers(COMPANY_SHEET)

    # name-identical, case-insensitive matches, plus the explicit rename pairs
    lowered = {h.lower(): h for h in com_headers}
    pairs = [
        (h, SHARED_COLUMN_ALIASES.get(h) or lowered.get(h.lower()))
        for h in ind_headers
    ]
    com_by_norm = {h.strip().lower(): h for h in com_headers}
    pairs = [
        (a, com_by_norm[b.strip().lower()])
        for a, b in pairs if b and b.strip().lower() in com_by_norm
    ]
    assert len(pairs) >= 20, f"expected the bureau/policy overlap, got {len(pairs)}"

    for code in individual:
        for ind_col, com_col in pairs:
            assert individual[code][ind_col] == company[code][com_col], (
                f"{code}: '{ind_col}' vs '{com_col}' drifted between the entity "
                f"matrices: {individual[code][ind_col]!r} != {company[code][com_col]!r}"
            )


@pytest.mark.parametrize("column", sorted(INERT_COLUMNS))
def test_inert_columns_stay_uniform(column: str) -> None:
    """An inert column that starts discriminating needs a rule, not silence."""
    sheet = INDIVIDUAL_SHEET if column in _headers(INDIVIDUAL_SHEET) else COMPANY_SHEET
    banks = _read(sheet)
    values = {str(policy[column]).strip().lower() for policy in banks.values()}

    assert values == {"true"}, (
        f"'{column}' is no longer permissive for every bank ({values}); "
        "it now affects verdicts and needs a rejection path"
    )


def test_no_column_is_left_deferred() -> None:
    """DEFERRED_COLUMNS parks columns awaiting a policy decision. Every
    discriminating column now has a rule, so it should stay empty."""
    assert DEFERRED_COLUMNS == {}, f"still awaiting a decision: {sorted(DEFERRED_COLUMNS)}"


def test_every_partner_bank_is_selectable_in_the_wizard() -> None:
    """A BankCode with no ExistingBankOption is unreachable, not just hidden.

    The wizard picks the assessed bank from `banking.existingAccountBank`, and
    every private bank carries `requires_existing_account = True`. An omitted
    bank therefore fails REL-501 on every single submission -- its column in
    `bank_eligibility` reads false regardless of the applicant -- so the option
    set has to stay in step with the matrix.
    """
    unreachable = set(BANK_MATRIX_RULES) - {
        code.value for code in EXISTING_BANK_TO_BANK_CODE.values()
    }
    assert not unreachable, (
        f"{sorted(unreachable)} have policy rows but no ExistingBankOption; "
        "REL-501 would reject them on every submission"
    )


@pytest.mark.parametrize("bank", sorted(FORM_SELECTABLE))
def test_selecting_a_bank_assesses_that_bank(bank: str) -> None:
    """Each option routes the verdict to its own bank's policy."""
    form = OnboardingFormRequest.model_validate(build_form({**DEFAULTS, "selected": bank}))

    assert form.banking.selected_bank.value == bank
    assert form.to_engine_payload()["existing_account_bank"] == bank
